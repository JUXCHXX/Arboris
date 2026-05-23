"""
ARBORIS - Export Module
Exportacion a PDF, Excel e importacion desde CSV
"""
import os
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from .tree import ArborisTree, Node


PDF_ICON_REPLACEMENTS = {
    "✅": "[OK]",
    "✓": "[OK]",
    "✔": "[OK]",
    "►": "[EN CURSO]",
    "▸": "[SUBTAREA]",
    "✗": "[X]",
    "✘": "[X]",
    "❌": "[X]",
    "○": "[ ]",
    "•": "-",
    "📁": "[PROY]",
    "📝": "[NOTA]",
    "🔴": "[VENCIDA]",
}


def _find_pdf_font_bundle():
    candidate_dirs = [Path(__file__).resolve().parent.parent / "assets" / "fonts"]
    windir = os.environ.get("WINDIR")
    if windir:
        candidate_dirs.append(Path(windir) / "Fonts")
    candidate_dirs.extend([
        Path("/usr/share/fonts/truetype/dejavu"),
        Path("/usr/share/fonts/truetype/liberation2"),
        Path("/Library/Fonts"),
    ])

    bundles = [
        {"family": "ArborisSans", "regular": "DejaVuSans.ttf", "bold": "DejaVuSans-Bold.ttf", "italic": "DejaVuSans-Oblique.ttf"},
        {"family": "ArborisSans", "regular": "arial.ttf", "bold": "arialbd.ttf", "italic": "ariali.ttf"},
        {"family": "ArborisSans", "regular": "segoeui.ttf", "bold": "segoeuib.ttf", "italic": "segoeuii.ttf"},
    ]

    for base_dir in candidate_dirs:
        for bundle in bundles:
            regular = base_dir / bundle["regular"]
            bold = base_dir / bundle["bold"]
            italic = base_dir / bundle["italic"]
            if regular.exists() and bold.exists():
                return {
                    "family": bundle["family"],
                    "regular": regular,
                    "bold": bold,
                    "italic": italic if italic.exists() else regular,
                }
    return None


def _configure_pdf_fonts(pdf):
    bundle = _find_pdf_font_bundle()
    if not bundle:
        return "Helvetica", False

    pdf.add_font(bundle["family"], "", fname=str(bundle["regular"]))
    pdf.add_font(bundle["family"], "B", fname=str(bundle["bold"]))
    pdf.add_font(bundle["family"], "I", fname=str(bundle["italic"]))
    return bundle["family"], True


def _clean_pdf_text(value, allow_unicode: bool) -> str:
    text = unicodedata.normalize("NFKC", str(value or ""))
    for source, replacement in PDF_ICON_REPLACEMENTS.items():
        text = text.replace(source, replacement)

    cleaned = []
    for char in text:
        if unicodedata.category(char) in {"So", "Cs"}:
            continue
        cleaned.append(char)

    normalized = "".join(cleaned).replace("\r\n", "\n").replace("\r", "\n")
    if allow_unicode:
        return normalized
    return normalized.encode("latin-1", "replace").decode("latin-1")


# PDF EXPORT
def export_pdf(tree: "ArborisTree", output_path: str = "exports/arboris_report.pdf"):
    from fpdf import FPDF

    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    DARK = (13, 15, 20)
    ACCENT = (0, 229, 190)
    WHITE = (240, 244, 255)
    MUTED = (136, 146, 164)
    WARN = (255, 107, 107)
    GOLD = (255, 209, 102)

    stats = tree.stats()

    class PDF(FPDF):
        def header(self):
            self.set_fill_color(*DARK)
            self.rect(0, 0, 210, 297, "F")
            self.set_fill_color(*ACCENT)
            self.rect(0, 0, 210, 8, "F")
            self.set_xy(10, 10)
            self.set_font(font_family, "B", 16)
            self.set_text_color(*WHITE)
            self.cell(0, 8, clean_text("ARBORIS - Reporte de Proyectos"), ln=True)
            self.set_font(font_family, "", 9)
            self.set_text_color(*MUTED)
            self.cell(0, 5, clean_text(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"), ln=True)
            self.ln(4)

        def footer(self):
            self.set_y(-12)
            self.set_font(font_family, "", 8)
            self.set_text_color(*MUTED)
            self.cell(0, 5, clean_text(f"Pagina {self.page_no()}"), align="C")

    pdf = PDF()
    font_family, has_unicode_font = _configure_pdf_fonts(pdf)
    clean_text = lambda text: _clean_pdf_text(text, has_unicode_font)
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font(font_family, "B", 11)
    pdf.set_text_color(*ACCENT)
    pdf.cell(0, 7, clean_text("Resumen General"), ln=True)
    pdf.set_draw_color(*ACCENT)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(3)

    stat_items = [
        (f"Total de nodos: {stats['total']}", WHITE),
        (f"Completados: {stats['completed']} ({stats['completion_rate']}%)", ACCENT),
        (f"Pendientes: {stats['pending']}", GOLD),
        (f"Vencidos: {stats['overdue']}", WARN),
    ]
    pdf.set_font(font_family, "", 10)
    for text, color in stat_items:
        pdf.set_text_color(*color)
        pdf.cell(0, 6, clean_text(text), ln=True)
    pdf.ln(5)

    def render_node(node, depth=0):
        left_margin = 10 + (depth * 8)
        width = max(60, 195 - left_margin)
        priority_colors = {"alta": WARN, "media": GOLD, "baja": ACCENT}
        col = priority_colors.get(node.priority, WHITE)
        status_label = {
            "completado": "[OK]",
            "en progreso": "[EN CURSO]",
            "cancelado": "[X]",
            "pendiente": "[ ]",
        }.get(node.status, "[ ]")

        pdf.set_font(font_family, "B", 10)
        pdf.set_text_color(*col)
        pdf.set_x(left_margin)
        pdf.multi_cell(width, 6, clean_text(f"{status_label} [{node.node_type.upper()}] {node.title}"))

        pdf.set_font(font_family, "", 8)
        pdf.set_text_color(*MUTED)
        meta_parts = [f"Prioridad: {node.priority}"]
        if node.due_date:
            meta_parts.append(f"Vence: {node.due_date}")
        if node.tags:
            meta_parts.append(f"Tags: {', '.join(node.tags)}")
        pdf.set_x(left_margin)
        pdf.multi_cell(width, 5, clean_text(" | ".join(meta_parts)))

        if node.notes:
            pdf.set_font(font_family, "I", 8)
            pdf.set_text_color(100, 120, 150)
            pdf.set_x(left_margin)
            pdf.multi_cell(width, 5, clean_text(f"Notas: {node.notes[:200]}"))

        pdf.ln(1)
        for child in node.children:
            render_node(child, depth + 1)

    pdf.set_font(font_family, "B", 11)
    pdf.set_text_color(*ACCENT)
    pdf.cell(0, 7, clean_text("Arbol de Proyectos"), ln=True)
    pdf.set_draw_color(*ACCENT)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(3)

    for root in tree.roots:
        render_node(root)
        pdf.ln(3)

    pdf.output(output_path)
    return output_path


# EXCEL EXPORT
def export_excel(tree: "ArborisTree", output_path: str = "exports/arboris_report.xlsx"):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tareas y Proyectos"

    headers = ["ID", "Titulo", "Tipo", "Prioridad", "Estado", "Vence", "Tags", "Notas", "Creado"]
    header_fill = PatternFill("solid", fgColor="00E5BE")
    header_font = Font(bold=True, color="0D0F14")
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    priority_fills = {
        "alta": PatternFill("solid", fgColor="FF6B6B"),
        "media": PatternFill("solid", fgColor="FFD166"),
        "baja": PatternFill("solid", fgColor="06D6A0"),
    }
    status_fills = {
        "completado": PatternFill("solid", fgColor="1B4332"),
        "en progreso": PatternFill("solid", fgColor="1C2030"),
        "cancelado": PatternFill("solid", fgColor="2D1B1B"),
        "pendiente": PatternFill("solid", fgColor="1A1A2E"),
    }

    row = 2
    for node in tree.all_nodes():
        values = [
            node.id, node.title, node.node_type, node.priority,
            node.status, node.due_date or "", ", ".join(node.tags),
            node.notes, node.created_at,
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = Font(color="F0F4FF")
            cell.fill = PatternFill("solid", fgColor="1C2030")
            if col == 4:
                cell.fill = priority_fills.get(node.priority, cell.fill)
            if col == 5:
                cell.fill = status_fills.get(node.status, cell.fill)
        row += 1

    col_widths = [10, 40, 12, 12, 14, 12, 25, 40, 18]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    ws.freeze_panes = "A2"
    wb.save(output_path)
    return output_path


# CSV IMPORT
def import_from_csv(path: str, tree: "ArborisTree") -> int:
    """
    Importa tareas desde CSV. Columnas esperadas:
    title, type, priority, due_date, tags, notes, status
    """
    import csv
    from .tree import Node

    if not os.path.exists(path):
        return 0

    imported = 0
    with open(path, newline="", encoding="utf-8") as file:
        reader = csv.DictReader(file)
        for row in reader:
            try:
                tags = [tag.strip() for tag in row.get("tags", "").split(",")] if row.get("tags") else []
                node = Node(
                    title=row.get("title", "Sin titulo"),
                    node_type=row.get("type", "tarea"),
                    priority=row.get("priority", "media"),
                    due_date=row.get("due_date") or None,
                    tags=tags,
                    notes=row.get("notes", ""),
                    status=row.get("status", "pendiente"),
                )
                tree.add_root(node)
                imported += 1
            except Exception:
                continue
    return imported
